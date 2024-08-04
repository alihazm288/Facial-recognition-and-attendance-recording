import tkinter as tk
from tkinter import filedialog,messagebox as mess,ttk
from tkinter import *
from PIL import ImageDraw
from datetime import datetime
import pandas as pd
import customtkinter
from openpyxl import Workbook
import openpyxl,xlrd
import pathlib
from openpyxl import load_workbook
import tkinter
from PIL import Image
import tkinter.simpledialog as tsd
import cv2,os
import csv
import numpy as np
import getpass
import re
from cv2 import *
import face_recognition
from PIL import ImageTk, Image
from tkinter import Tk
import arabic_reshaper
from pynput.keyboard import Controller, Key
from concurrent.futures import ThreadPoolExecutor
import tkinter.messagebox as messagebox


camera = None
is_capturing = False


def resize_frame(image, scale_percent):
    width = int(image.shape[1] * scale_percent / 100)
    height = int(image.shape[0] * scale_percent / 100)
    dim = (width, height)
    resized_frame = cv2.resize(image, dim, interpolation=cv2.INTER_AREA)
    return resized_frame


def start_capture():
    global is_capturing
    is_capturing = True


def stop_capture():
    global is_capturing
    is_capturing = False



def close_camera():
    global camera
    if camera is not None:

        camera.release()



def Quit():
    root.destroy()


    
def extract_name_and_id(image_name):
    name_1 = re.sub(r'[^a-zA-Z\s]+', '', image_name)
    iD1 = re.sub(r'[^0-9]+', '', image_name)
    return name_1, iD1

path = 'images'
images = []
classNames = []
personsList = os.listdir(path)


for cl in personsList:
    curPersonn = cv2.imread(f'{path}/{cl}')
    images.append(curPersonn)
    classNames.append(os.path.splitext(cl)[0])


def findEncodeings(image):
    encodeList = []
    for image in images:
        image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(image)[0]
        encodeList.append(encode)
    return encodeList
encodeListKnown = findEncodeings(images)
print('Encoding Complete.')




def capture_image():
    name = name_entry.get()
    ID = ID_entry.get()
    if name.strip() == "" or ID.strip() == "":
        messagebox.showerror("Error", "Please enter a name and ID")

    else:
        
        img_counter=0
        global camera
        camera = cv2.VideoCapture(0)
        while (camera.isOpened()):
        
            return_value, image = camera.read()
    
            if return_value==True:


                resized_frame = resize_frame(image, 31)
                cv2image = cv2.cvtColor(resized_frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(cv2image)
                imgtk = ImageTk.PhotoImage(image=img)
                panel.config(image=imgtk)
                panel.update()
                img2 = Image.open(image_path1)
                img2 = img2.resize((200, 148)) 
                imgtk1 = ImageTk.PhotoImage(image=img2)
                panel.imgtk = imgtk1
                panel.config(image=imgtk1)


            if is_capturing:
                img_name= "images\ " + name +"." +ID+'.'+ str()+ "jpg".format()
                cv2.imwrite(img_name,image)
                
                print('Photo Save'.format(img_counter))

                img2 = Image.open(image_path1)
                img2 = img2.resize((200, 148)) 
                imgtk1 = ImageTk.PhotoImage(image=img2)
                panel.imgtk = imgtk1
                panel.config(image=imgtk1)
                stop_capture()  
                panel.update()
                camera.release()
                
                
                name = name_entry.get()
                ID = ID_entry.get()

                data = {'Name': [name], 'ID': [ID]}
                df = pd.DataFrame(data)

                if os.path.isfile('data.csv'):
                    df.to_csv('data.csv', mode='a', header=False, index=False)
                else:
                    df.to_csv('data.csv', index=False)
                encodeListKnown = findEncodeings(images)
                print('Encoding Complete.')


def recognize_person():
    global camera
    
    camera = cv2.VideoCapture(0)
    current_time1 = datetime.now().strftime("%Y-%m-%d")
    if os.path.exists("Attendance_" + current_time1 +".xlsx"):
        wb = load_workbook("Attendance_" + current_time1 +".xlsx")
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = 'Name'
        sheet['B1'] = 'ID'
        sheet['C1'] = 'Time'
        

    while True:
        return_value, image = camera.read()
        imgS = cv2.resize(image, (0,0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
        faceCurentFrame = face_recognition.face_locations(imgS)
        encodeCurentFrame = face_recognition.face_encodings(imgS, faceCurentFrame)

        for encodeface, faceLoc in zip(encodeCurentFrame, faceCurentFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeface)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeface)
            matchIndex = np.argmin(faceDis)
            if matches[matchIndex]:
                name = classNames[matchIndex].upper()
                
                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1*4, x2*4, y2*4, x1*4
                cv2.rectangle(image, (x1, y1), (x2, y2), (0,255,0), 2)
                cv2.rectangle(image, (x1,y2-35), (x2,y2), (0,255,0), cv2.FILLED)
                image_name = name
                name1, iD = extract_name_and_id(image_name)
                cv2.putText(image, name1,(x1+6, y2-6), cv2.FONT_HERSHEY_COMPLEX, 1, (255,255,255), 2)        
                name1 ,iD = extract_name_and_id(image_name)
                ID = str(iD)
                name=name1

            else:
                Id = 'Unknown'
                bb = str(Id)
                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1*4, x2*4, y2*4, x1*4
                cv2.rectangle(image, (x1, y1), (x2, y2), (0,0,255), 2)
                cv2.rectangle(image, (x1,y2-35), (x2,y2), (0,0,255), cv2.FILLED)
                cv2.putText(image, str(bb), (x1+6, y2-6), cv2.FONT_HERSHEY_COMPLEX, 1, (255,255,255), 2)                        
                

            
            
            current_time = datetime.now().strftime('%I:%M:%p')
            found = False
            for row in sheet.iter_rows(min_row=2, max_col=2):
               if row[0].value == name:
                   found = True
                   sheet.cell(row=row[1].row, column=3).value = current_time
                   break

            if not found:
                sheet.append([name,ID , current_time])
        cv2.namedWindow('Face Recognition', cv2.WINDOW_NORMAL)
        
        cv2.imshow('Face Recognition', image)
        current_time1 = datetime.now().strftime("%Y-%m-%d")
        wb.save("Attendance_" + current_time1 +".xlsx")
        
        if cv2.waitKey(1) & 0xFF == ord('q'):
           break
        

    camera.release()
    cv2.destroyAllWindows()

   




customtkinter.set_appearance_mode("dark-blue")
customtkinter.set_default_color_theme("green")


root = customtkinter.CTk()
root.title('facial recognition attendance system')
root.geometry("600x660")

frame = customtkinter.CTkFrame(root)
frame.pack(pady=20, padx=30, fill="both", expand=True)

image_folder = "image_interface"
image_path1 = os.path.join(os.getcwd(), image_folder, "persons.png")


label2 = customtkinter.CTkLabel(frame, text="NEW REGISTER" )
label2.place(x=225, y=20 )

name_entry = customtkinter.CTkEntry(frame,placeholder_text="Name"
,justify=tkinter.CENTER,
font=customtkinter.CTkFont(size=15, family="Lalezar -Regular"),
text_color="#f5f5f5" )
name_entry.place(x=60, y=70 )


label3 = customtkinter.CTkLabel(frame, text="ATTENDANCE")
label3.place(x=230,y=320)

current_time = datetime.now().strftime('%I:%M:%p')
label4 = customtkinter.CTkLabel(frame, text=current_time)
label4.place(x=450,y=20)

ID_entry = customtkinter.CTkEntry(frame,placeholder_text="ID"
,justify=tkinter.CENTER,
font=customtkinter.CTkFont(size=15, family="Lalezar -Regular"),
text_color="#f5f5f5" )
ID_entry.place(x=60,y=150)

panel = Frame(frame ,bd=2, relief=SUNKEN)
panel.place(x=450, y=80)
image = PhotoImage(file=image_path1)
panel =Label(panel ,image=image)
panel.pack()


Open_camera = customtkinter.CTkButton(frame, text="Open camera", command=capture_image ,height = 26,width = 20)
Open_camera.place(x=28,y=220)



capture_button = customtkinter.CTkButton(frame,text="capture_image",command=start_capture)
capture_button.place(x=355, y=220)


recognize_button = customtkinter.CTkButton(frame, text="Face Recognition", command=recognize_person)
recognize_button.place(x=200,y=410)

recognize_button = customtkinter.CTkButton(frame, text="close camera", command=close_camera,height = 26,width = 20 )
recognize_button.place(x=138,y=220)

button = customtkinter.CTkButton(frame, text="Quit", command=Quit,font=customtkinter.CTkFont(size=20, family="Lalezar -Regular"),fg_color='#f70717')
button.place(x=200,y=520)



root.mainloop()


